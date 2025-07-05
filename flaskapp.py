from yolodetection import detect_webcam, detect_video, detect_photo, analyze_safety_stats, get_class_colors, print_log, colorize_source
from flask import Flask, Response, request, session, render_template, flash, redirect, url_for, send_file, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from collections import Counter
from datetime import datetime
from fpdf import FPDF
import subprocess
import atexit
import os
import cv2

app = Flask(__name__)
app.secret_key = 'smart_ppe_detection_system'
app.config['UPLOAD_FOLDER'] = 'static/files'


# Cleanup static folder on exit
def cleanup_static_files():
    try:
        subprocess.run(["cleanup_static.bat"], check=True)
        print("\033[1;92mstatic folder has been cleaned!\033[0m", flush=True)
    except Exception as e:
        print(f"\033[1;91mFailed to clean static/files folder: {e}\033[0m", flush=True)

atexit.register(cleanup_static_files)


# Cleanup results folder on exit
def cleanup_results_files():
    try:
        subprocess.run(["cleanup_results.bat"], check=True)
        print("\033[1;92mresults folder has been cleaned!\033[0m", flush=True)
    except Exception as e:
        print(f"\033[1;91mFailed to clean results folder: {e}\033[0m", flush=True)

atexit.register(cleanup_results_files)


# Home Page
@app.route('/')
def home():
    return render_template('home.html')


# Home Page background photo and favicon
@app.route('/assets/<path:filename>')
def assets_file(filename):
    return send_from_directory('assets', filename)


# Webcam Page
@app.route('/webcam')
def webcam():
    return render_template('webcam.html')


# Webcam Feed
@app.route('/webcam_feed')
def webcam_feed():
    print_log("📷 Live webcam detection started.", "info")
    return Response(detect_webcam(), mimetype='multipart/x-mixed-replace; boundary=frame')


# Video Upload Page
@app.route('/video', methods=['GET', 'POST'])
def video():
    if request.method == 'POST':
        if 'video' not in request.files:
            print_log("Video upload failed: no file field found in form.", "error")
            return "No video file uploaded!", 400

        video = request.files['video']
        if video.filename == '':
            print_log("Video upload failed: empty file submitted.", "error")
            return "Empty file submitted!", 400

        os.makedirs("static/video", exist_ok=True)
        filename = secure_filename(video.filename)
        path = os.path.join("static/video", filename)
        video.save(path)

        print_log(f"🎞️  Video uploaded: \033[91m{filename}\033[0m", "success")
        session['video_path'] = path
        return redirect(url_for('video', uploaded='true'))
    
    if request.args.get('uploaded') != 'true':
        session.pop('video_path', None)
    return render_template('video.html')


# Process and Show Uploaded Video
@app.route('/video_feed')
def video_feed():
    path = session.get('video_path', None)
    if path and os.path.exists(path):
        print_log("🎥 Video detection started.", "info")
        return Response(detect_video(path), mimetype='multipart/x-mixed-replace; boundary=frame')
    return "Video not found!", 404


# Photo Upload Page
@app.route('/photo', methods=['GET', 'POST'])
def photo():
    if request.method == 'POST':
        file = request.files["image"]
        os.makedirs("static/photo", exist_ok=True)
        os.makedirs("static/photo_results", exist_ok=True)

        filename = secure_filename(file.filename)
        upload_path = os.path.join("static/photo", filename)
        file.save(upload_path)

        print_log(f"🖼️  Photo uploaded: \033[91m{filename}\033[0m", "success")
        print_log("🖼️  Photo detection started.", "info")
        class_counts, result_path, safety_score, predicted_env = detect_photo(upload_path)

        return render_template(
            "photo.html",
            image_file=result_path,
            class_counts=class_counts,
            class_colors=get_class_colors(),
            safety_score=safety_score,
            predicted_env=predicted_env
        )
    return render_template("photo.html", image_file=None)


# Gallery Page
@app.route('/gallery')
def gallery():
    video_dir = 'static/video_results'
    photo_dir = 'static/photo_results'
    original_video_dir = 'static/video'
    original_photo_dir = 'static/photo'
    thumb_dir = 'static/thumbs'

    def save_video_thumbnail(video_path, thumb_path):
        cap = cv2.VideoCapture(video_path)
        success, frame = cap.read()
        if success:
            os.makedirs(os.path.dirname(thumb_path), exist_ok=True)
            cv2.imwrite(thumb_path, frame)
        cap.release()

    def read_logs():
        logs = {}
        try:
            with open("results/log.txt", "r", encoding="utf-8") as file:
                for line in file:
                    parts = line.strip().split()
                    if len(parts) >= 7:
                        date = parts[0] + ' ' + parts[1]
                        source = parts[2].strip("[]")
                        env = parts[3].strip("[]")
                        filename = parts[4]
                        score = parts[-1]
                        logs[filename] = {
                            'score': score,
                            'env': env,
                            'date': date
                        }
        except FileNotFoundError:
            pass
        return logs

    logs = read_logs()
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    
    video_items = []
    photo_items = []

    if os.path.exists(video_dir):
        for filename in os.listdir(video_dir):
            if filename.endswith('.mp4'):
                name = os.path.splitext(filename)[0]
                video_path = os.path.join(video_dir, filename)
                original_basename = filename.replace("detected_", "")
                original_path = os.path.join(original_video_dir, original_basename)
                thumb_path = os.path.join(thumb_dir, f"{name}.jpg")

                if not os.path.exists(thumb_path):
                    save_video_thumbnail(original_path, thumb_path)
                meta = logs.get(filename, {})

                video_items.append({
                    'original_video': f"/{original_path}",
                    'detected_video': f"/{video_path}?v={timestamp}",
                    'thumbnail_image': f"/{thumb_path}",
                    'score': meta.get('score', 'N/A'),
                    'env': meta.get('env', 'N/A'),
                    'date': meta.get('date', 'Unknown')
                })

    if os.path.exists(photo_dir):
        for filename in os.listdir(photo_dir):
            if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                original_name = filename.replace("detected_", "")
                original_path = os.path.join(original_photo_dir, original_name)
                detected_path = os.path.join(photo_dir, filename)
                meta = logs.get(filename, {})

                photo_items.append({
                    'original_photo': f"/{original_path}",
                    'detected_photo': f"/{detected_path}",
                    'score': meta.get('score', 'N/A'),
                    'env': meta.get('env', 'N/A'),
                    'date': meta.get('date', 'Unknown')
                })

    return render_template('gallery.html', videos=video_items, photos=photo_items)


# Gallery upload/download messages in terminal logs
@app.route('/log-gallery-action', methods=['POST'])
def log_gallery_action():
    data = request.get_json()
    action = data.get('action', '').upper()
    if action == "DOWNLOAD_VIDEO":
        print_log("⬇️  \033[93mDetected Video\033[0m successfully downloaded from gallery.", "success")
    elif action == "DOWNLOAD_PHOTO":
        print_log("⬇️  \033[91mDetected Photo\033[0m successfully downloaded from gallery.", "success")
    elif action == "UPLOAD_VIDEO":
        print_log("🞧 \033[93mVideo upload page\033[0m accessed from gallery.", "info")
    elif action == "UPLOAD_PHOTO":
        print_log("🞧 \033[91mPhoto upload page\033[0m accessed from gallery.", "info")
    else:
        print_log("Unknown gallery action received.", "warn")
    return '', 204


# Detection Log Page
@app.route('/log')
def log():
    source_filter = request.args.get('source', 'ALL').upper()
    environment_filter = request.args.get('environment', 'ALL').upper()
    score_filter = request.args.get('score', 'ALL').upper()
    page = int(request.args.get('page', 1))
    per_page = 20

    logs = []

    log_file = 'results/log.txt'
    if os.path.exists(log_file):
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        if source_filter != 'ALL':
           lines = [line for line in lines if f"[{source_filter.upper()}]" in line.upper()]

        if environment_filter != 'ALL':
           lines = [line for line in lines if f"[{environment_filter.upper()}]" in line.upper()]

        if score_filter != 'ALL':
            def in_score_range(score):
                try:
                    score = float(score.replace('%', '').strip())
                    if score_filter == "LOW":
                        return 0 <= score <= 35
                    elif score_filter == "MEDIUM":
                        return 36 <= score <= 70
                    elif score_filter == "HIGH":
                        return 71 <= score <= 100
                except Exception:
                    return False
            lines = [line for line in lines if in_score_range(line.strip().split()[-1])]

        total_logs = len(lines)
        start = (page - 1) * per_page
        end = start + per_page
        logs = lines[start:end]

        total_pages = (total_logs + per_page - 1) // per_page

        max_display_pages = 9
        half_range = max_display_pages // 2
        if total_pages <= max_display_pages:
            page_range = list(range(1, total_pages + 1))
        else:
            start_page = max(1, page - half_range)
            end_page = min(total_pages, start_page + max_display_pages - 1)
            start_page = max(1, end_page - max_display_pages + 1)
            page_range = list(range(start_page, end_page + 1))

    else:
        total_pages = 1
        page_range = [1]

    return render_template(
        'log.html',
        logs=logs,
        current_page=page,
        total_pages=total_pages,
        page_range=page_range
    )


# Excel Download
@app.route('/download-excel')
def download_excel():
    source_filter = (request.args.get('source') or 'ALL').strip().upper()
    environment_filter = (request.args.get('environment') or 'ALL').strip().upper()
    score_filter = (request.args.get('score') or 'ALL').strip().upper()
    log_file = 'results/log.txt'

    if not os.path.exists(log_file):
        print_log("⚠️  Excel export failed: log file not found.", "error")
        flash("No detection data available to export as Excel.", "warning")
        return redirect(url_for('log'))

    with open(log_file, 'r') as f:
        lines = f.readlines()

    if source_filter != 'ALL':
        lines = [line for line in lines if f"[{source_filter}]" in line.upper()]

    if environment_filter != 'ALL':
        lines = [line for line in lines if f"[{environment_filter}]" in line.upper()]

    if score_filter != 'ALL':
        def in_score_range(score):
            try:
                score = float(score.replace('%', '').strip())
                if score_filter == "LOW":
                    return 0 <= score <= 35
                elif score_filter == "MEDIUM":
                    return 36 <= score <= 70
                elif score_filter == "HIGH":
                    return 71 <= score <= 100
            except Exception:
                return False
        lines = [line for line in lines if in_score_range(line.strip().split()[-1])]
    
    if not lines:
        flash("No matching detection records available to export as Excel.", "warning")
        return redirect(url_for('log', source=request.args.get('source'), environment=request.args.get('environment'), score=request.args.get('score')))

    wb = Workbook()
    ws = wb.active
    ws.title = "Detection Log"

    headers = ["Timestamp", "Source", "Environment", "Filename", "Class", "Confidence", "Score"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")

    for line in lines:
        parts = line.strip().split()
        if len(parts) >= 7:
            timestamp = parts[0] + ' ' + parts[1]
            source = parts[2].strip("[]")
            environment = parts[3].strip("[]")
            filename = parts[4]
            class_name = parts[5]
            confidence = parts[6]
            score = parts[7]
            ws.append([timestamp, source, environment, filename, class_name, confidence, score])

    for column_cells in ws.columns:
        col_letter = column_cells[0].column_letter

        header = ws[f"{col_letter}1"].value
        if header == "Source":
            ws.column_dimensions[col_letter].width = 15
        elif header == "Environment":
            ws.column_dimensions[col_letter].width = 17
        elif header == "Filename":
            ws.column_dimensions[col_letter].width = 28
        elif header == "Class":
            ws.column_dimensions[col_letter].width = 15
        elif header == "Score":
            ws.column_dimensions[col_letter].width = 12
        else:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[col_letter].width = length + 2

        for cell in column_cells:
            cell.alignment = Alignment(horizontal="center")

        if source_filter == 'ALL' and environment_filter == 'ALL' and score_filter == 'ALL':
            output_path = "results/all_log_export.xlsx"
        else:
            output_path = f"results/{source_filter.lower()}_{environment_filter.lower()}_{score_filter.lower()}_log_export.xlsx"
    try:
        wb.save(output_path)
        print_log(f"📊 Excel report for {colorize_source(source_filter)} | Env: \033[95m{environment_filter}\033[0m | Score: \033[92m{score_filter}\033[0m exported successfully.", "success")
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        print_log(f"⚠️ Excel export failed during saving or sending file: {e}", "error")
        flash("An unexpected error occurred while exporting Excel file.", "danger")
        return redirect(url_for('log'))


# PDF class with page numbers
class MyPDF(FPDF):
    def footer(self):
        self.set_y(-15)
        self.set_font("DejaVu", size=9)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align='C')


# PDF Report
@app.route('/download-pdf')
def download_pdf():
    source_filter = (request.args.get('source') or 'ALL').strip().upper()
    environment_filter = (request.args.get('environment') or 'ALL').strip().upper()
    score_filter = (request.args.get('score') or 'ALL').strip().upper()

    log_file = 'results/log.txt'
    font_path = os.path.join(os.path.dirname(__file__), "fonts", "DejaVuSans.ttf")

    if not os.path.exists(log_file):
        print_log("⚠️  PDF export failed: log file not found.", "error")
        flash("No detection data available to export as PDF.", "warning")
        return redirect(url_for('log'))

    if not os.path.exists(font_path):
        return "Font file not found!", 500

    with open(log_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if source_filter != 'ALL':
        lines = [line for line in lines if f"[{source_filter}]" in line.upper()]

    if environment_filter != 'ALL':
        lines = [line for line in lines if f"[{environment_filter}]" in line.upper()]

    if score_filter != 'ALL':
        def in_score_range(score):
            try:
                score = float(score.replace('%', '').strip())
                if score_filter == "LOW":
                    return 0 <= score <= 35
                elif score_filter == "MEDIUM":
                    return 36 <= score <= 70
                elif score_filter == "HIGH":
                    return 71 <= score <= 100
            except Exception:
                return False
        lines = [line for line in lines if in_score_range(line.strip().split()[-1])]

    if not lines:
        flash("No matching detection records available to export as PDF.", "warning")
        return redirect(url_for('log', source=request.args.get('source'), environment=request.args.get('environment'), score=request.args.get('score')))

    pdf = MyPDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    pdf.add_font("DejaVu", "", font_path, uni=True)
    pdf.add_font("DejaVu", "B", font_path, uni=True)

    timestamp_w = 35
    source_w = 23
    environment_w = 25
    filename_w = 43
    class_w = 23
    conf_w = 23
    score_w = 20
    total_w = timestamp_w + source_w + environment_w + filename_w + class_w + conf_w + score_w

    pdf.set_font("DejaVu", size=12, style="B")
    pdf.cell(0, 10, "Smart PPE Detection Report", ln=True, align='C')
    pdf.ln(10)

    pdf.set_font("DejaVu", size=9, style="B")
    pdf.set_x((210 - total_w) / 2)
    pdf.cell(timestamp_w, 10, "Timestamp", border=1, align='C')
    pdf.cell(source_w, 10, "Source", border=1, align='C')
    pdf.cell(environment_w, 10, "Environment", border=1, align='C')
    pdf.cell(filename_w, 10, "Filename", border=1, align='C')
    pdf.cell(class_w, 10, "Class", border=1, align='C')
    pdf.cell(conf_w, 10, "Confidence", border=1, align='C')
    pdf.cell(score_w, 10, "Score", border=1, align='C')
    pdf.ln()

    pdf.set_font("DejaVu", size=8)

    for line in lines:
        parts = line.strip().split()
        if len(parts) >= 7:
            timestamp = parts[0] + ' ' + parts[1]
            source = parts[2].strip("[]")
            environment = parts[3].strip("[]")
            filename = parts[4]
            class_name = parts[5]
            confidence = parts[6]
            score = parts[7]

            pdf.set_x((210 - total_w) / 2)
            pdf.cell(timestamp_w, 10, timestamp, border=1)
            pdf.cell(source_w, 10, source, border=1)
            pdf.cell(environment_w, 10, environment, border=1)
            pdf.cell(filename_w, 10, filename, border=1)
            pdf.cell(class_w, 10, class_name, border=1)
            pdf.cell(conf_w, 10, confidence, border=1)
            pdf.cell(score_w, 10, score, border=1)
            pdf.ln()

    if source_filter == 'ALL' and environment_filter == 'ALL' and score_filter == 'ALL':
        output_path = "results/all_log_report.pdf"
    else:
        output_path = f"results/{source_filter.lower()}_{environment_filter.lower()}_{score_filter.lower()}_log_report.pdf"
    
    try:
        pdf.output(output_path)
        print_log(f"🧾 PDF report for {colorize_source(source_filter)} | Env: \033[95m{environment_filter}\033[0m | Score: \033[92m{score_filter}\033[0m exported successfully.", "success")
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        print_log(f"⚠️ PDF export failed during saving or sending file: {e}", "error")
        flash("An unexpected error occurred while exporting PDF file.", "danger")
        return redirect(url_for('log'))


# Chart Page
@app.route('/chart')
def chart():
    log_file = 'results/log.txt'
    class_counts_video = {}
    class_counts_webcam = {}
    class_counts_photo = {}

    stats_video = None
    stats_webcam = None
    stats_photo = None

    if os.path.exists(log_file):
        with open(log_file, 'r') as f:
            lines = f.readlines()
        
        video_classes = [line.strip().split()[5] for line in lines if "[VIDEO]" in line]
        webcam_classes = [line.strip().split()[5] for line in lines if "[WEBCAM]" in line]
        photo_classes = [line.strip().split()[5] for line in lines if "[PHOTO]" in line]

        class_counts_video = dict(Counter(video_classes)) if video_classes else {}
        class_counts_webcam = dict(Counter(webcam_classes)) if webcam_classes else {}
        class_counts_photo = dict(Counter(photo_classes)) if photo_classes else {}

        stats_video = analyze_safety_stats(lines, "VIDEO") if video_classes else None
        stats_webcam = analyze_safety_stats(lines, "WEBCAM") if webcam_classes else None
        stats_photo = analyze_safety_stats(lines, "PHOTO") if photo_classes else None

    if not class_counts_video and not class_counts_webcam and not class_counts_photo:
        flash("No detection data available. Please run a detection first.", "danger")

    return render_template(
        'chart.html',
        class_counts_video=class_counts_video,
        class_counts_webcam=class_counts_webcam,
        class_counts_photo=class_counts_photo,
        class_colors=get_class_colors(),
        stats_video=stats_video,
        stats_webcam=stats_webcam,
        stats_photo=stats_photo
    )


# Chart PNG download messages in terminal logs
@app.route('/log-chart-download', methods=['POST'])
def log_chart_download():
    data = request.get_json()
    source = data.get('source', '').upper()
    if source == "VIDEO":
        print_log("🎞️  PNG \033[95mfor\033[0m \033[93m'Video Analysis Chart'\033[0m downloaded successfully.", "success")
    elif source == "WEBCAM":
        print_log("📸 PNG \033[95mfor\033[0m \033[96m'Webcam Overview Chart'\033[0m downloaded successfully.", "success")
    elif source == "PHOTO":
        print_log("🖼️  PNG \033[95mfor\033[0m \033[91m'Photo Overview Chart'\033[0m downloaded successfully.", "success")
    else:
        print_log("Unknown chart download source.", "warn")
    return '', 204


# Start the Flask App
if __name__ == '__main__':
    print_log("🚀 Flask application started successfully.", "success")
    # Disables the automatic reloader to prevent the app from running twice in debug mode
    app.run(debug=True, use_reloader=False)